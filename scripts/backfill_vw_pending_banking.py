#!/usr/bin/env python3
"""
scripts/backfill_vw_pending_banking.py

Stage 1 remediation for the VW + Auto Pedigree pending backlog
surfaced by the diagnostic in PR #29.

The diagnostic showed 205 pending records across 5 templates with
significant banking-data gaps. The Revio API does NOT let us cancel,
delete, or force-activate pending BTCs (status is read-only when
pending; subscription_billing_scheduled_on is read-only on the BTC;
no DELETE endpoint for individual BTCs). The one PATCH path the API
DOES expose is `/clients/{id}/`, which is enough to repair the
banking columns on the underlying Client. If Revio's internal AVS
loop holds pending records back specifically because of bad banking,
fixing that data may let some auto-promote to active.

WHAT THIS SCRIPT DOES:
  1. Read SALES tab (the existing dashboard sheet) — phone-keyed
     index of bank_account + bank_code per VW customer.
  2. Walk the 5 target templates, list pending BTCs, fetch each
     Client (cached by client_id).
  3. For VW pending records (VW Premium HELP Single R89 + Family
     R159 = 118 records):
       - Probe Client.bank_account + bank_code with the same shape
         checks the diagnostic used.
       - If both already OK → skip (nothing to fix).
       - Else look up the customer in SALES by normalised phone.
       - If SALES has BETTER values (probe = ok) → PATCH the Client.
       - Lock PATCH payload to {bank_account, bank_code} only —
         belt-and-braces guard against any future code path
         accidentally including status / personal_code / etc.
  4. For Auto Pedigree pending records (87 records across 3
     templates) → emit a CSV/Excel sheet for JD's manual upload.
     The Auto Ped banking source lives in
     /Users/.../OneDrive/.../Auto Ped/Data which the runner can't
     reach, so manual handover is the only path here.

WHAT THIS SCRIPT EXPLICITLY DOES NOT DO:
  - Touch personal_code (the existing convert_account_expiry pattern
    is `"VW-" + Policy Number`; whether to switch that to SA ID is
    a separate decision).
  - Set subscription_billing_scheduled_on (read-only on the BTC).
  - Cancel duplicate BTCs (no API endpoint exists; would need a
    Revio support ticket).
  - Touch any record that isn't in {pending} status across the
    target templates.

POPIA — bank account / personal code values appear in the Excel
output (it's the deliverable). stdout logs mask PII (first-4 +
last-4 chars only). Email always goes to JD only — no production
fan-out switch — because this is a remediation tool with PATCH
side effects and the deliverable is JD-owned.

ENV:
  REVIO_API_KEY              required
  GOOGLE_SHEETS_CREDENTIALS  service-account JSON for SALES read +
                             audit row append
  EMAIL_SENDER               Gmail sender
  EMAIL_PASSWORD             Gmail app password
  DRY_RUN                    "true" (default) → no PATCHes, full
                             plan emailed to JD; "false" → live
                             PATCH against Revio
"""

import json
import logging
import mimetypes
import os
import smtplib
import sys
import time
from datetime import datetime, timezone
from email.message import EmailMessage
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from email_template import RunSummary, build_run_summary_email  # noqa: E402
from revio_subscription import (  # noqa: E402
    REVIO_API_BASE_URL,
    _do_request_with_retry,
    _get_headers,
    normalise_phone,
)
from silence_existing_revio_subscribers import (  # noqa: E402
    list_clients_for_template,
    list_subscription_templates,
)

logger = logging.getLogger(__name__)

DRY_RUN = os.environ.get("DRY_RUN", "true").lower() == "true"
JD_RECIPIENT = "jd@projecthelp.co.za"

# ─── Targets ─────────────────────────────────────────────────────────────────
VW_TEMPLATE_TITLES = frozenset({
    "VW Premium HELP Single R89",
    "VW Premium HELP Family R159",
})
AUTO_PED_TEMPLATE_TITLES = frozenset({
    "Auto Ped Embedded Family",
    "Auto Pedigree Family R159",
    "Auto Pedigree R89",
})
TARGET_TEMPLATE_TITLES = VW_TEMPLATE_TITLES | AUTO_PED_TEMPLATE_TITLES

PENDING_STATUS = "pending"

# ─── SALES sheet identity (mirrors sync_sales_to_sheets.py) ──────────────────
SHEET_ID  = "1nzDkzva7wZO0lDFBDctNQdqxvOU-uexyUkxmex6xGgs"
SALES_TAB = "SALES"

SALES_PHONE_COL    = "Mobile Number (VW/Audi Campaign 1)"
SALES_BANK_ACC_COL = "Bank Account Number (VW/Audi)"
SALES_BANK_CODE_COL = "Branch Code (VW/Audi Campaign 1)"

# ─── PATCH safety guard ─────────────────────────────────────────────────────
# The PATCH payload is locked to exactly these two fields. Any other key
# in the payload raises before the network call — defensive safety net
# against any future code path accidentally injecting `status`,
# `personal_code`, etc. Mirrors the pattern from
# silence_existing_revio_subscribers.py:_assert_patch_payload_safe.
ALLOWED_PATCH_FIELDS = frozenset({"bank_account", "bank_code"})

# ─── ZA-shape probes (must match analyse_vw_autoped_pending probes) ─────────
ZA_BANK_ACCOUNT_MIN_DIGITS = 9
ZA_BANK_ACCOUNT_MAX_DIGITS = 11
ZA_BANK_CODE_DIGITS        = 6

SHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
AUDIT_TAB     = "BACKFILL_AUDIT"
AUDIT_HEADER  = [
    "Run Timestamp (UTC ISO)", "Mode", "VW Pending Examined",
    "VW Already OK", "VW Patched (or Would Patch)", "VW No SALES Match",
    "VW Sales Data Bad", "Auto Ped Pending Examined",
    "Errors", "Triggered By", "Output Filename",
]


# ─── Pure helpers ────────────────────────────────────────────────────────────
def _digits_only(s):
    return "".join(c for c in str(s or "") if c.isdigit())


def probe_bank_account(value):
    s = str(value or "").strip()
    if not s:
        return "blank"
    digits = _digits_only(s)
    if not digits:
        return "non_numeric"
    if len(digits) < ZA_BANK_ACCOUNT_MIN_DIGITS:
        return "too_short"
    if len(digits) > ZA_BANK_ACCOUNT_MAX_DIGITS:
        return "too_long"
    return "ok"


def probe_bank_code(value):
    """Trims to digits + length-checks against ZA's 6-digit branch
    code. Returns one of: blank, wrong_length, ok."""
    s = str(value or "").strip()
    if not s:
        return "blank"
    digits = _digits_only(s)
    if len(digits) != ZA_BANK_CODE_DIGITS:
        return "wrong_length"
    return "ok"


def classify_template(title):
    """Pure: bucket a template title by remediation path. Returns
    "vw" / "auto_ped" / "other". Used by the main loop to fan out
    each pending BTC to the correct handler."""
    if title in VW_TEMPLATE_TITLES:
        return "vw"
    if title in AUTO_PED_TEMPLATE_TITLES:
        return "auto_ped"
    return "other"


def build_sales_phone_index(sales_rows):
    """Pure: sales_rows is a list of {column_name: value} dicts (one
    per SALES data row). Returns {normalised_phone: row_dict}.

    Empty/blank phones are dropped. Duplicate phones use last-write-
    wins (later SALES rows overwrite earlier ones — SALES is sorted
    by Created Time per sync_sales_to_sheets.py:254, so "last wins"
    means the most recent sign-up's banking is what we use).
    """
    out = {}
    for row in sales_rows:
        raw_phone = row.get(SALES_PHONE_COL)
        n = normalise_phone(raw_phone)
        if not n:
            continue
        out[n] = row
    return out


def compute_patch_payload(client_bank_acc, client_bank_code,
                          sales_bank_acc, sales_bank_code):
    """Pure: decide what to PATCH on the Client. Returns the PATCH
    payload dict (possibly empty) — caller skips PATCH if empty.

    Conservative rules:
      - Only patch a field if the client's current value FAILS the
        probe (blank/non_numeric/too_short/too_long/wrong_length).
      - And SALES's value PASSES the probe (no point overwriting bad
        with bad). SALES values that fail their own probe are
        treated as "no improvement available" and left alone.
      - Never touch a field that's already OK on the client — even
        if SALES has a different OK value, we don't second-guess
        live data with potentially-stale SALES data.

    The two-field independence is deliberate: it's normal for a
    record to have a clean bank_account but a 5-digit bank_code, or
    vice versa, and patching just the broken one is the right call.
    """
    payload = {}
    if probe_bank_account(client_bank_acc) != "ok":
        if probe_bank_account(sales_bank_acc) == "ok":
            payload["bank_account"] = str(sales_bank_acc).strip()
    if probe_bank_code(client_bank_code) != "ok":
        if probe_bank_code(sales_bank_code) == "ok":
            payload["bank_code"] = str(sales_bank_code).strip()
    return payload


def assert_patch_payload_safe(payload):
    """Pre-flight: PATCH payload must contain ONLY the two banking
    keys. Mirrors silence_existing_revio_subscribers._assert_patch_
    payload_safe — belt-and-braces against any future code path
    accidentally injecting `status` / `personal_code` / etc."""
    extra = set(payload.keys()) - ALLOWED_PATCH_FIELDS
    if extra:
        raise RuntimeError(
            f"PATCH payload contains unexpected keys: {sorted(extra)}. "
            f"Only banking fields are permitted to prevent accidental "
            f"modification of other Client state. "
            f"Allowed: {sorted(ALLOWED_PATCH_FIELDS)}."
        )


def _mask_pii(value, head=4, tail=4):
    s = "" if value is None else str(value)
    if len(s) <= head + tail:
        return s
    return f"{s[:head]}…{s[-tail:]}"


# ─── Sheets read ─────────────────────────────────────────────────────────────
def _get_sheets_service():
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds_json = os.environ.get("GOOGLE_SHEETS_CREDENTIALS")
    if not creds_json:
        raise RuntimeError("GOOGLE_SHEETS_CREDENTIALS not set")
    creds = Credentials.from_service_account_info(
        json.loads(creds_json), scopes=SHEETS_SCOPES,
    )
    return build("sheets", "v4", credentials=creds)


def read_sales_tab(svc):
    """Returns list of dicts (one per SALES data row), keyed by
    header column name. UNFORMATTED_VALUE matches the pattern used
    by convert_account_expiry / sync_sales_to_sheets so phone /
    account fields don't arrive as scientific notation."""
    res = svc.spreadsheets().values().get(
        spreadsheetId=SHEET_ID,
        range=f"{SALES_TAB}!A1:ZZ",
        valueRenderOption="UNFORMATTED_VALUE",
    ).execute()
    rows = res.get("values", [])
    if not rows:
        return []
    header = rows[0]
    n = len(header)
    out = []
    for r in rows[1:]:
        padded = list(r) + [""] * (n - len(r))
        out.append(dict(zip(header, padded)))
    return out


# ─── Revio reads ─────────────────────────────────────────────────────────────
def fetch_client(client_id, cache):
    if not client_id:
        return None
    if client_id in cache:
        return cache[client_id]
    url = REVIO_API_BASE_URL + f"/clients/{client_id}/"
    r = _do_request_with_retry("get", url, headers=_get_headers(), timeout=30)
    if r is None or r.status_code != 200:
        cache[client_id] = None
        return None
    try:
        data = r.json()
    except ValueError:
        cache[client_id] = None
        return None
    cache[client_id] = data
    return data


# ─── Revio write ─────────────────────────────────────────────────────────────
def patch_client_banking(client_id, payload):
    """PATCH /clients/{id}/ with a JSON body containing only the
    banking fields. Returns True on success, False on failure (logged
    but not raised — one bad PATCH must not abort the whole loop).
    Asserts payload safety before the call."""
    assert_patch_payload_safe(payload)
    url = REVIO_API_BASE_URL + f"/clients/{client_id}/"
    r = _do_request_with_retry(
        "patch", url, headers=_get_headers(), json=payload, timeout=30,
    )
    if r is None:
        logger.error("PATCH /clients/%s/ — transport error after retries",
                     client_id)
        return False
    if r.status_code not in (200, 202, 204):
        body = (r.text or "")[:300]
        logger.error("PATCH /clients/%s/ returned %d: %s",
                     client_id, r.status_code, body)
        return False
    return True


# ─── Excel writer ────────────────────────────────────────────────────────────
def write_workbook(path, vw_actions, auto_ped_records, summary):
    """Three sheets:
      Summary — totals + plan
      VW Patches — what was/would be patched (and what was skipped + why)
      Auto Ped Manual — full record list for JD's manual upload
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, summary)

    ws_vw = wb.create_sheet("VW Patches")
    _write_vw_actions(ws_vw, vw_actions, get_column_letter, Font, PatternFill,
                      Alignment)

    ws_ap = wb.create_sheet("Auto Ped Manual")
    _write_auto_ped(ws_ap, auto_ped_records, get_column_letter, Font)

    wb.save(path)


def _write_summary(ws, s):
    from openpyxl.styles import Font
    bold = Font(bold=True)
    title = Font(bold=True, size=14)
    ws["A1"] = "VW Pending Banking Backfill — Plan / Outcome"
    ws["A1"].font = title
    ws["A2"] = (f"Generated: "
                f"{datetime.now(timezone.utc).isoformat(timespec='seconds')} "
                f"({'DRY RUN' if DRY_RUN else 'PRODUCTION'})")
    row = 4
    for label, value in [
        ("VW pending examined",            s["vw_examined"]),
        ("VW already OK (skipped)",        s["vw_already_ok"]),
        ("VW patched / would-patch",       s["vw_patched"]),
        ("VW dedup-skipped (sibling BTC)", s.get("vw_dedup_skipped", 0)),
        ("VW no SALES match (manual)",     s["vw_no_match"]),
        ("VW SALES data also bad",         s["vw_sales_bad"]),
        ("VW PATCH errors",                s["vw_errors"]),
        ("Auto Ped pending examined",      s["auto_ped_examined"]),
        ("Auto Ped manual list rows",      s["auto_ped_listed"]),
    ]:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16


def _write_vw_actions(ws, actions, col_letter, font, fill, align):
    bold = font(bold=True)
    fills = {
        "patched":          fill("solid", fgColor="C8E6C9"),  # green
        "would_patch":      fill("solid", fgColor="C8E6C9"),
        "already_ok":       fill("solid", fgColor="E0E0E0"),  # grey
        "no_sales_match":   fill("solid", fgColor="FFE0B2"),  # amber
        "sales_data_bad":   fill("solid", fgColor="FFCDD2"),  # red
        "patch_error":      fill("solid", fgColor="FFCDD2"),
    }
    headers = ["Outcome", "Template", "Client ID", "Personal Code",
               "Phone (Client)", "Phone (Normalised)",
               "Current Bank Account", "Current Bank Code",
               "SALES Bank Account", "SALES Bank Code",
               "Patch Bank Account", "Patch Bank Code", "Note"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
    ws.freeze_panes = "A2"
    for ri, a in enumerate(actions, 2):
        cells = [
            a["outcome"], a["template"], a["client_id"], a["personal_code"],
            a["phone_raw"], a["phone_norm"],
            a["client_bank_acc"], a["client_bank_code"],
            a["sales_bank_acc"], a["sales_bank_code"],
            a["patch_bank_acc"], a["patch_bank_code"], a["note"],
        ]
        for ci, v in enumerate(cells, 1):
            cell = ws.cell(row=ri, column=ci, value="" if v is None else v)
            if ci == 1 and a["outcome"] in fills:
                cell.fill = fills[a["outcome"]]
    for i, h in enumerate(headers, 1):
        max_len = len(h)
        for a in actions:
            v = a.get(["outcome", "template", "client_id", "personal_code",
                       "phone_raw", "phone_norm", "client_bank_acc",
                       "client_bank_code", "sales_bank_acc", "sales_bank_code",
                       "patch_bank_acc", "patch_bank_code", "note"][i - 1])
            if v and len(str(v)) > max_len:
                max_len = len(str(v))
        ws.column_dimensions[col_letter(i)].width = min(max_len + 2, 50)


def _write_auto_ped(ws, records, col_letter, font):
    bold = font(bold=True)
    headers = ["Template", "BT Client ID", "Client ID", "Personal Code",
               "Full Name", "Phone", "Email",
               "Current Bank Account", "Bank Account Status",
               "Current Bank Code", "Bank Code Status",
               "Days Since Created"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h).font = bold
    ws.freeze_panes = "A2"
    for ri, r in enumerate(records, 2):
        for ci, v in enumerate([
            r.get("template"), r.get("bt_client_id"), r.get("client_id"),
            r.get("personal_code"), r.get("full_name"), r.get("phone"),
            r.get("email"),
            r.get("bank_account"), r.get("bank_account_status"),
            r.get("bank_code"), r.get("bank_code_status"),
            r.get("days_since_created"),
        ], 1):
            ws.cell(row=ri, column=ci, value="" if v is None else v)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[col_letter(i)].width = min(len(h) + 8, 40)


# ─── Audit row ───────────────────────────────────────────────────────────────
def _ensure_audit_tab(svc):
    meta = svc.spreadsheets().get(
        spreadsheetId=SHEET_ID, fields="sheets(properties(title))",
    ).execute()
    titles = {s["properties"]["title"] for s in meta.get("sheets", [])}
    if AUDIT_TAB in titles:
        return
    svc.spreadsheets().batchUpdate(
        spreadsheetId=SHEET_ID,
        body={"requests": [{"addSheet": {"properties": {"title": AUDIT_TAB}}}]},
    ).execute()
    svc.spreadsheets().values().update(
        spreadsheetId=SHEET_ID, range=f"{AUDIT_TAB}!A1",
        valueInputOption="USER_ENTERED",
        body={"values": [AUDIT_HEADER]},
    ).execute()
    logger.info("Created %s tab + header in master sheet", AUDIT_TAB)


def append_audit_row(svc, row):
    try:
        _ensure_audit_tab(svc)
        svc.spreadsheets().values().append(
            spreadsheetId=SHEET_ID, range=f"{AUDIT_TAB}!A1",
            valueInputOption="USER_ENTERED",
            insertDataOption="INSERT_ROWS",
            body={"values": [row]},
        ).execute()
        logger.info("Appended %s row", AUDIT_TAB)
    except Exception as e:
        logger.error("Failed to append %s row: %s", AUDIT_TAB, e)


# ─── Email ───────────────────────────────────────────────────────────────────
def send_email(run_date, summary, attachment_path, error_summary,
               duration_seconds):
    sender = os.environ.get("EMAIL_SENDER")
    pwd    = os.environ.get("EMAIL_PASSWORD")
    if not sender or not pwd:
        logger.warning("EMAIL_SENDER / EMAIL_PASSWORD not set — skipping email")
        return

    if error_summary:
        outcome  = "failure"
        headline = "Backfill failed"
        para = f"Backfill failed: {error_summary}. Manual investigation needed."
    elif summary["vw_patched"] == 0 and not error_summary:
        outcome  = "noop"
        headline = "No VW patches needed / available"
        para = (f"Examined {summary['vw_examined']} VW pending record(s). "
                f"{summary['vw_already_ok']} already had clean banking; "
                f"{summary['vw_no_match']} had no SALES match; "
                f"{summary['vw_sales_bad']} had SALES data that was also bad. "
                f"Nothing to patch.")
    else:
        outcome  = "success"
        verb = "would be patched" if DRY_RUN else "patched"
        headline = f"{summary['vw_patched']} VW Client(s) {verb}"
        para = (
            f"Examined {summary['vw_examined']} VW pending record(s). "
            f"{summary['vw_patched']} {verb} from SALES. "
            f"{summary['vw_already_ok']} already OK, "
            f"{summary['vw_no_match']} no SALES match, "
            f"{summary['vw_sales_bad']} SALES data bad, "
            f"{summary['vw_errors']} error(s). "
            f"Auto Pedigree: {summary['auto_ped_listed']} record(s) listed "
            f"in the workbook for JD's manual upload (no auto-fix path "
            f"for Auto Ped — banking source is local-only). POPIA: this "
            f"workbook contains personal data — handle per policy. Do "
            f"not forward externally."
        )

    rs = RunSummary(
        workflow_name="Revio — Backfill VW Pending Banking",
        run_date=run_date,
        mode="dry_run" if DRY_RUN else "production",
        outcome=outcome,
        headline=headline,
        summary_paragraph=para,
        numbers={
            "VW pending examined":          summary["vw_examined"],
            "VW already OK":                summary["vw_already_ok"],
            ("VW would-patch" if DRY_RUN else "VW patched"):
                                            summary["vw_patched"],
            "VW dedup-skipped (sibling)":   summary.get("vw_dedup_skipped", 0),
            "VW no SALES match":            summary["vw_no_match"],
            "VW SALES data bad":            summary["vw_sales_bad"],
            "VW errors":                    summary["vw_errors"],
            "Auto Ped pending examined":    summary["auto_ped_examined"],
            "Auto Ped manual list rows":    summary["auto_ped_listed"],
        },
        duration_seconds=duration_seconds,
        attachments_note=(f"Excel attached: {Path(attachment_path).name}"
                          if attachment_path else None),
    )
    subject, html_body = build_run_summary_email(rs)

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"]    = sender
    msg["To"]      = JD_RECIPIENT
    msg.set_content("HTML email — see HTML part for the run summary.")
    msg.add_alternative(html_body, subtype="html")

    if attachment_path and Path(attachment_path).exists():
        ctype, _ = mimetypes.guess_type(attachment_path)
        if not ctype:
            ctype = ("application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet")
        maintype, subtype = ctype.split("/", 1)
        with open(attachment_path, "rb") as f:
            msg.add_attachment(f.read(), maintype=maintype, subtype=subtype,
                               filename=Path(attachment_path).name)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(sender, pwd)
        smtp.send_message(msg)
    logger.info("Summary email sent → %s%s",
                JD_RECIPIENT, " (DRY RUN)" if DRY_RUN else "")


# ─── Main ────────────────────────────────────────────────────────────────────
def _triggered_by():
    return os.environ.get("GITHUB_ACTOR") or "manual"


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )
    run_date = datetime.now(timezone.utc)
    started  = time.monotonic()
    logger.info("=" * 60)
    logger.info("REVIO — VW PENDING BANKING BACKFILL")
    logger.info("Run date    : %s", run_date.isoformat(timespec="seconds"))
    logger.info("Dry run     : %s", DRY_RUN)
    logger.info("Triggered by: %s", _triggered_by())
    logger.info("=" * 60)

    output_path = Path("/tmp") / (
        f"vw_pending_banking_backfill_"
        f"{run_date.date().isoformat()}.xlsx"
    )
    summary = {
        "vw_examined": 0, "vw_already_ok": 0, "vw_patched": 0,
        "vw_no_match": 0, "vw_sales_bad": 0,
        "vw_dedup_skipped": 0,  # client_id already PATCHed earlier in run
        "vw_errors": 0,
        "auto_ped_examined": 0, "auto_ped_listed": 0,
    }
    vw_actions = []
    auto_ped_records = []

    try:
        # ── Pre-flight: SALES read ───────────────────────────────────────────
        svc = _get_sheets_service()
        sales_rows = read_sales_tab(svc)
        sales_index = build_sales_phone_index(sales_rows)
        logger.info("SALES tab: %d data rows, %d unique phone keys",
                    len(sales_rows), len(sales_index))

        # ── Enumerate target templates + pending BTCs ────────────────────────
        templates = list_subscription_templates()
        targeted = [
            t for t in templates
            if (t.get("title") or t.get("name") or "") in TARGET_TEMPLATE_TITLES
        ]
        if len(targeted) != len(TARGET_TEMPLATE_TITLES):
            found = {t.get("title") or t.get("name") for t in targeted}
            missing = TARGET_TEMPLATE_TITLES - found
            logger.warning("Target template(s) not found: %s", sorted(missing))

        # Collect pending BTCs (template_title, btc) for both buckets.
        vw_pending = []
        ap_pending = []
        for tmpl in targeted:
            tid    = tmpl.get("id") or tmpl.get("uuid")
            ttitle = tmpl.get("title") or tmpl.get("name")
            btcs = list_clients_for_template(tid)
            for b in btcs:
                if b.get("status") != PENDING_STATUS:
                    continue
                bucket = classify_template(ttitle)
                if bucket == "vw":
                    vw_pending.append((ttitle, b))
                elif bucket == "auto_ped":
                    ap_pending.append((ttitle, b))
        logger.info("Pending BTCs found: %d VW, %d Auto Ped",
                    len(vw_pending), len(ap_pending))

        # ── Fetch unique Clients (VW + Auto Ped) ─────────────────────────────
        client_cache = {}
        unique_cids = (
            {b.get("client_id") for _, b in vw_pending if b.get("client_id")}
            | {b.get("client_id") for _, b in ap_pending if b.get("client_id")}
        )
        logger.info("Fetching %d unique Client record(s)", len(unique_cids))
        for n, cid in enumerate(unique_cids, 1):
            fetch_client(cid, client_cache)
            if n % 50 == 0:
                logger.info("  …fetched %d / %d", n, len(unique_cids))

        # ── VW remediation pass ──────────────────────────────────────────────
        # Client-level dedupe: the diagnostic showed avg ~3.5 BTCs per
        # client (one client had 10 dupes on Auto Pedigree R89). For
        # the same client_id appearing multiple times in vw_pending we
        # only PATCH the first time — subsequent BTC rows are listed
        # in the workbook with outcome=already_patched_this_run so the
        # sheet is honest, but we don't waste 3 PATCHes / hit the same
        # Client thrice with the same payload.
        patched_this_run = set()  # client_ids
        for ttitle, btc in vw_pending:
            summary["vw_examined"] += 1
            cid = btc.get("client_id")
            client = client_cache.get(cid) or {}
            phone_raw = client.get("phone")
            phone_norm = normalise_phone(phone_raw)
            client_acc  = client.get("bank_account") or ""
            client_code = client.get("bank_code") or ""

            base_action = {
                "template":          ttitle,
                "client_id":         cid,
                "personal_code":     client.get("personal_code") or "",
                "phone_raw":         phone_raw or "",
                "phone_norm":        phone_norm,
                "client_bank_acc":   client_acc,
                "client_bank_code":  client_code,
                "sales_bank_acc":    "",
                "sales_bank_code":   "",
                "patch_bank_acc":    "",
                "patch_bank_code":   "",
                "note":              "",
            }

            if cid and cid in patched_this_run:
                # This client_id already received its (would-)PATCH on
                # an earlier BTC row in this run. Don't double-patch
                # the same Client; just record the BTC for visibility.
                summary["vw_dedup_skipped"] += 1
                vw_actions.append({**base_action,
                                   "outcome": "already_patched_this_run",
                                   "note": ("Client already PATCHed earlier "
                                            "in this run via another BTC")})
                continue

            # Already-OK short-circuit
            if (probe_bank_account(client_acc) == "ok"
                    and probe_bank_code(client_code) == "ok"):
                summary["vw_already_ok"] += 1
                vw_actions.append({**base_action,
                                   "outcome": "already_ok",
                                   "note": "Both fields pass probe; "
                                           "no patch needed"})
                continue

            sales_match = sales_index.get(phone_norm) if phone_norm else None
            if not sales_match:
                summary["vw_no_match"] += 1
                vw_actions.append({**base_action,
                                   "outcome": "no_sales_match",
                                   "note": ("No SALES row for this phone — "
                                            "manual lookup required")})
                continue

            sales_acc  = sales_match.get(SALES_BANK_ACC_COL) or ""
            sales_code = sales_match.get(SALES_BANK_CODE_COL) or ""
            base_action["sales_bank_acc"]  = sales_acc
            base_action["sales_bank_code"] = sales_code

            payload = compute_patch_payload(client_acc, client_code,
                                            sales_acc, sales_code)
            if not payload:
                summary["vw_sales_bad"] += 1
                vw_actions.append({**base_action,
                                   "outcome": "sales_data_bad",
                                   "note": ("SALES data also fails probe / "
                                            "no improvement available")})
                continue

            base_action["patch_bank_acc"]  = payload.get("bank_account", "")
            base_action["patch_bank_code"] = payload.get("bank_code", "")

            logger.info(
                "  %s VW client_id=%s personal_code=%s phone=%s "
                "patch=%s bank_acc(masked)=%s",
                ("would_patch" if DRY_RUN else "patch"),
                cid, client.get("personal_code") or "∅",
                _mask_pii(phone_norm),
                list(payload.keys()),
                _mask_pii(payload.get("bank_account") or client_acc),
            )

            if DRY_RUN:
                summary["vw_patched"] += 1
                if cid:
                    patched_this_run.add(cid)
                vw_actions.append({**base_action, "outcome": "would_patch",
                                   "note": "DRY RUN — not sent to Revio"})
                continue

            ok = patch_client_banking(cid, payload)
            if ok:
                summary["vw_patched"] += 1
                if cid:
                    patched_this_run.add(cid)
                vw_actions.append({**base_action, "outcome": "patched",
                                   "note": "PATCH /clients/ → 200"})
            else:
                summary["vw_errors"] += 1
                vw_actions.append({**base_action, "outcome": "patch_error",
                                   "note": "PATCH failed — see log"})

        # ── Auto Ped pass — list only, no auto-fix ───────────────────────────
        for ttitle, btc in ap_pending:
            summary["auto_ped_examined"] += 1
            cid = btc.get("client_id")
            client = client_cache.get(cid) or {}
            client_acc  = client.get("bank_account") or ""
            client_code = client.get("bank_code") or ""
            try:
                created = btc.get("created_on")
                if isinstance(created, (int, float)) and not isinstance(
                        created, bool):
                    days = (run_date.date()
                            - datetime.fromtimestamp(
                                float(created), tz=timezone.utc).date()).days
                else:
                    days = ""
            except Exception:
                days = ""
            auto_ped_records.append({
                "template":              ttitle,
                "bt_client_id":          btc.get("id") or "",
                "client_id":             cid,
                "personal_code":         client.get("personal_code") or "",
                "full_name":             client.get("full_name") or "",
                "phone":                 normalise_phone(client.get("phone")),
                "email":                 client.get("email") or "",
                "bank_account":          client_acc,
                "bank_account_status":   probe_bank_account(client_acc),
                "bank_code":             client_code,
                "bank_code_status":      probe_bank_code(client_code),
                "days_since_created":    days,
            })
            summary["auto_ped_listed"] += 1

        # ── Workbook + audit + email ─────────────────────────────────────────
        # Workflow-log self-sufficiency: dump every bucket count so
        # the run is fully describable from the log alone, without
        # cracking open the Excel or the email.
        logger.info("─" * 60)
        logger.info("VW SUMMARY (DRY RUN=%s)", DRY_RUN)
        logger.info("  examined                : %d", summary["vw_examined"])
        logger.info("  already_ok              : %d", summary["vw_already_ok"])
        logger.info("  %s          : %d",
                    "would_patch" if DRY_RUN else "patched    ",
                    summary["vw_patched"])
        logger.info("  no_sales_match          : %d", summary["vw_no_match"])
        logger.info("  sales_data_bad          : %d", summary["vw_sales_bad"])
        logger.info("  dedup_skipped (sibling) : %d",
                    summary["vw_dedup_skipped"])
        logger.info("  errors                  : %d", summary["vw_errors"])
        logger.info("AUTO PED SUMMARY")
        logger.info("  examined                : %d",
                    summary["auto_ped_examined"])
        logger.info("  listed (manual)         : %d",
                    summary["auto_ped_listed"])
        logger.info("─" * 60)

        write_workbook(output_path, vw_actions, auto_ped_records, summary)
        logger.info("Wrote workbook → %s", output_path)

        append_audit_row(svc, [
            run_date.isoformat(timespec="seconds"),
            "dry_run" if DRY_RUN else "production",
            summary["vw_examined"], summary["vw_already_ok"],
            summary["vw_patched"], summary["vw_no_match"],
            summary["vw_sales_bad"], summary["auto_ped_examined"],
            summary["vw_errors"], _triggered_by(), output_path.name,
        ])

        send_email(run_date, summary, str(output_path),
                   error_summary="",
                   duration_seconds=time.monotonic() - started)

    except Exception as e:
        logger.exception("Backfill failed: %s", e)
        try:
            send_email(run_date, summary,
                       str(output_path) if output_path.exists() else None,
                       error_summary=str(e),
                       duration_seconds=time.monotonic() - started)
        except Exception as e2:
            logger.error("Summary email also failed: %s", e2)
        sys.exit(1)

    logger.info("=" * 60)
    logger.info("DONE")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
