#!/usr/bin/env python3
"""
scripts/dump_non_active_revio_subscribers.py

One-off CRM dump of every non-active Revio subscriber (status in
{subscription_paused, inactive, pending}) so the CS team can re-engage
them onto R89 single / R159 family.

Background — PR #26 silence backfill surfaced 5,259 non-active
subscribers across all subscription templates:
  paused   : 2,838  (voluntarily paused subscriptions)
  pending  : 2,285  (added to template but never fully launched)
  inactive :   136  (API-internal state, read-only)

JD wants a multi-sheet Excel workbook containing the full CRM record for
each, attached to one summary email. Tashlyn + Tsakane work the call
list. The Excel is NOT committed to git — it lives in /tmp during the
run, then exists only as the email attachment + the DUMP_AUDIT row in
the VW Reporting Master Book.

WHAT IT DOES:
  1. List every subscription billing template.
  2. List every BillingTemplateClient (BTC) per template.
  3. Bucket BTCs by status — only {subscription_paused, inactive,
     pending} make it into the dump; active are noise (they are paying)
     and unknown statuses are logged + counted but skipped.
  4. Build the cross-ref index: for every ACTIVE BTC, fetch the Client,
     normalise the phone, key into {phone: [template_title, ...]}.
     Protects CS from cold-calling someone who is paying on another
     template under a different sign-up.
  5. For every non-active BTC, fetch the Client (cached by client_id),
     join into a flat row, look up the active phone index, and emit one
     workbook row.
  6. Write a 4-sheet Excel workbook to /tmp:
       Summary  — totals, top templates, currently-active-elsewhere counts
       Paused   — sorted by Days Since Created ASC (freshest first)
       Pending  — sorted by Days Since Created ASC
       Inactive — sorted by Days Since Created ASC
  7. Append one row to the DUMP_AUDIT tab in the VW Reporting Master
     Book (creates the tab + header if missing).
  8. Send a RunSummary email with the workbook attached.
       Production: jd, tashlyn, tsakane.
       Dry-run:    jd only (still attaches workbook so JD can sanity-
                   check format before broader distribution).

POPIA — the workbook contains personal identifiers (ID number, bank
account, phone, email, address). Logging masks PII (first-4 + last-4
chars only) and the email body carries a "do not forward externally"
reminder. The persistent record is the DUMP_AUDIT row, not the file.

ENV:
  REVIO_API_KEY              required
  REVIO_BRAND_ID             optional — scope to one brand (matches
                             silence_existing_revio_subscribers.py)
  GOOGLE_SHEETS_CREDENTIALS  service-account JSON for DUMP_AUDIT append
  EMAIL_SENDER               Gmail sender for summary email
  EMAIL_PASSWORD             Gmail app password
  EMAIL_RECIPIENT            comma-separated production recipient list
  DRY_RUN                    "true" (default) → JD only, "false" → full
"""

import json
import logging
import mimetypes
import os
import smtplib
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from email.message import EmailMessage
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from email_template import RunSummary, build_run_summary_email  # noqa: E402
from revio_subscription import (  # noqa: E402
    REVIO_API_BASE_URL,
    _do_request_with_retry,
    _get_headers,
    normalise_phone,
)
from silence_existing_revio_subscribers import (  # noqa: E402
    list_clients_for_template,
    list_subscription_templates,
)

logger = logging.getLogger(__name__)

DRY_RUN = os.environ.get("DRY_RUN", "true").lower() == "true"
DRY_RUN_RECIPIENT = "jd@projecthelp.co.za"

# Hardcoded per spec — POPIA-sensitive deliverable, locked to the three
# people who run the call list. Deliberately NOT sourced from the
# shared EMAIL_RECIPIENT secret so a future widening of that secret
# (e.g. to add the analytics distribution list to the silence workflow)
# can never accidentally fan this dump out to a wider audience.
PRODUCTION_RECIPIENTS = (
    "jd@projecthelp.co.za",
    "tashlyn@projecthelp.co.za",
    "tsakane@projecthelp.co.za",
)

# ─── Status taxonomy ─────────────────────────────────────────────────────────
# Internal correctness uses the live API value (subscription_paused);
# human-readable Excel/email copy uses "Paused" for readability. The
# split between API-string and friendly-label lives only in this file —
# do not propagate "Paused" as a status comparison anywhere.
STATUS_PAUSED   = "subscription_paused"
STATUS_INACTIVE = "inactive"
STATUS_PENDING  = "pending"
STATUS_ACTIVE   = "active"

NON_ACTIVE_STATUSES = frozenset({STATUS_PAUSED, STATUS_INACTIVE, STATUS_PENDING})

STATUS_LABEL = {
    STATUS_PAUSED:   "Paused",
    STATUS_INACTIVE: "Inactive",
    STATUS_PENDING:  "Pending",
}

BUCKET_PAUSED       = "paused"
BUCKET_PENDING      = "pending"
BUCKET_INACTIVE     = "inactive"
BUCKET_SKIP_ACTIVE  = "SKIP_active"
BUCKET_SKIP_UNKNOWN = "SKIP_unknown"

STATUS_TO_BUCKET = {
    STATUS_PAUSED:   BUCKET_PAUSED,
    STATUS_PENDING:  BUCKET_PENDING,
    STATUS_INACTIVE: BUCKET_INACTIVE,
    STATUS_ACTIVE:   BUCKET_SKIP_ACTIVE,
}

# ─── Excel column contract ───────────────────────────────────────────────────
# Order matters — each downstream sheet writer iterates this list.
COLUMNS = [
    "Status",
    "Template Title",
    "Template ID",
    "BT Client ID",
    "Client ID",
    "Personal Code",
    "Full Name",
    "Phone",
    "Email",
    "Bank Account",
    "Bank Code",
    "Street Address",
    "City",
    "Zip Code",
    "Invoice Reference",
    "Date Created",
    "Days Since Created",
    "Last Scheduled Debit",
    "Currently Active Elsewhere",
    "Active Elsewhere Templates",
]

# ─── Audit log target ────────────────────────────────────────────────────────
SHEET_ID  = "1nzDkzva7wZO0lDFBDctNQdqxvOU-uexyUkxmex6xGgs"
AUDIT_TAB = "DUMP_AUDIT"
AUDIT_HEADER = [
    "Run Timestamp (UTC ISO)",
    "Mode",
    "Total Examined",
    "Paused Count",
    "Pending Count",
    "Inactive Count",
    "Currently Active Elsewhere Count",
    "Triggered By",
    "Output Filename",
]
SHEETS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


# ─── Pure helpers ────────────────────────────────────────────────────────────
def classify_non_active(btc):
    """Pure: bucket a BTC by its status field.

    Returns one of: "paused", "pending", "inactive", "SKIP_active",
    "SKIP_unknown". Active subscribers are explicitly recognised (and
    skipped here — they are processed by the cross-ref pass instead, not
    dumped). Anything not in the four documented values is SKIP_unknown
    so API drift surfaces in logs rather than silently joining a bucket.
    """
    return STATUS_TO_BUCKET.get(btc.get("status"), BUCKET_SKIP_UNKNOWN)


def build_active_phone_index(active_records):
    """Pure: given an iterable of {"phone": str, "template_title": str}
    records (one per active BTC, with phone already normalised), return
    {normalised_phone: [template_title, ...]}.

    Empty / missing phones are dropped — keying on "" would collapse
    every phoneless active subscriber into one bucket and produce false
    "active elsewhere" matches. Template-title duplicates within a phone
    are deduped + sorted for stable output.
    """
    by_phone = defaultdict(set)
    for rec in active_records:
        phone = (rec.get("phone") or "").strip()
        if not phone:
            continue
        title = rec.get("template_title") or ""
        by_phone[phone].add(title)
    return {p: sorted(titles) for p, titles in by_phone.items()}


def annotate_with_active_elsewhere(record, active_phone_index):
    """Pure: mutate `record` in-place adding the two cross-ref columns.

    Reads `record["Phone"]` (already normalised by caller). If empty,
    or not present in the index, marks N + blank list. Otherwise marks
    Y and joins the matched template titles with "; ".
    """
    phone = (record.get("Phone") or "").strip()
    if not phone:
        record["Currently Active Elsewhere"]   = "N"
        record["Active Elsewhere Templates"]   = ""
        return record
    titles = active_phone_index.get(phone)
    if not titles:
        record["Currently Active Elsewhere"]   = "N"
        record["Active Elsewhere Templates"]   = ""
    else:
        record["Currently Active Elsewhere"]   = "Y"
        record["Active Elsewhere Templates"]   = "; ".join(titles)
    return record


def build_summary_sheet_data(dump_records):
    """Pure: collapse the full dump into the structure the Summary sheet
    needs. Returns:

      {
        "total":                  int,
        "by_status":              {"Paused": n, "Pending": n, "Inactive": n},
        "by_template":            [(title, n), ...] sorted DESC by n,
        "top_templates":          [(title, n), ...] top 10 by total
                                  non-active count,
        "top_templates_by_paused":
            [(title, n), ...] top 5 by paused-only count (drives the
            email next-steps line — the freshest re-engagement leads
            are the recently-paused ones),
        "active_elsewhere_total": int,
        "active_elsewhere_by_status":
            {"Paused": n, "Pending": n, "Inactive": n},
      }

    All counts are derived from `dump_records` only — no API calls. The
    function is the single source of truth for the Summary sheet so the
    sheet writer stays a dumb cell-painter.
    """
    by_status = Counter()
    by_template = Counter()
    by_template_paused = Counter()
    active_elsewhere_total = 0
    active_elsewhere_by_status = Counter()

    for r in dump_records:
        s = r.get("Status") or ""
        title = r.get("Template Title") or ""
        by_status[s] += 1
        by_template[title] += 1
        if s == "Paused":
            by_template_paused[title] += 1
        if r.get("Currently Active Elsewhere") == "Y":
            active_elsewhere_total += 1
            active_elsewhere_by_status[s] += 1

    by_template_sorted = by_template.most_common()
    return {
        "total":                  len(dump_records),
        "by_status": {
            "Paused":   by_status.get("Paused", 0),
            "Pending":  by_status.get("Pending", 0),
            "Inactive": by_status.get("Inactive", 0),
        },
        "by_template":             by_template_sorted,
        "top_templates":           by_template_sorted[:10],
        "top_templates_by_paused": by_template_paused.most_common(5),
        "active_elsewhere_total":  active_elsewhere_total,
        "active_elsewhere_by_status": {
            "Paused":   active_elsewhere_by_status.get("Paused", 0),
            "Pending":  active_elsewhere_by_status.get("Pending", 0),
            "Inactive": active_elsewhere_by_status.get("Inactive", 0),
        },
    }


def _mask_pii(value, head=4, tail=4):
    """Return a debug-safe string: first `head` + last `tail` chars only.
    Empty / short values come back unchanged ("" → ""). Used in stdout
    logging — the Excel dump itself carries the full value because
    that's literally the deliverable."""
    s = "" if value is None else str(value)
    if len(s) <= head + tail:
        return s
    return f"{s[:head]}…{s[-tail:]}"


def _days_since(iso_str, today):
    """ISO-ish date string → integer days since today, or "" if blank /
    unparseable. Tolerant of trailing 'Z' and time component."""
    s = (iso_str or "").strip()
    if not s:
        return ""
    try:
        # Accept "2025-09-19", "2025-09-19T10:30:00", "2025-09-19T10:30:00Z"
        d = datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        return ""
    return (today - d).days


def _iso_date_only(iso_str):
    """Strip time/tz from an ISO date so the Excel cell is a clean date.
    Returns "" for blank/unparseable inputs."""
    s = (iso_str or "").strip()
    if not s:
        return ""
    try:
        d = datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        return s  # fall back to raw — better partial info than dropped data
    return d.isoformat()


# ─── Revio reads (Client fetch + caching) ────────────────────────────────────
def fetch_client(client_id, cache):
    """GET /clients/{id}/ with in-run cache by client_id.

    Returns the Client dict on success, or None on persistent failure
    (logged, never raises — one bad fetch must not abort the whole 5K-
    row dump). The cache survives the run; a None result is also cached
    so we don't re-attempt a known-bad id.
    """
    if not client_id:
        return None
    if client_id in cache:
        return cache[client_id]

    url = REVIO_API_BASE_URL + f"/clients/{client_id}/"
    r = _do_request_with_retry("get", url, headers=_get_headers(), timeout=30)
    if r is None or r.status_code != 200:
        sc = r.status_code if r is not None else "transport-error"
        body = (r.text if r is not None else "")[:200]
        logger.warning("GET /clients/%s/ failed: %s %s", client_id, sc, body)
        cache[client_id] = None
        return None
    try:
        data = r.json()
    except ValueError:
        logger.warning("GET /clients/%s/ returned non-JSON body", client_id)
        cache[client_id] = None
        return None
    cache[client_id] = data
    return data


# ─── Record assembly ─────────────────────────────────────────────────────────
def build_dump_record(btc, client, template_title, template_id, today):
    """Pure: join a non-active BTC with its underlying Client into one
    flat dict keyed by COLUMNS. Cross-ref columns are populated later
    by annotate_with_active_elsewhere.

    `client` may be None if the GET /clients/{id}/ failed — in that case
    the personal-detail columns come back blank but the BTC-side columns
    (status, ids, dates, invoice ref) are still populated, so CS at
    least knows which BTC failed to enrich.
    """
    client = client or {}
    status_label = STATUS_LABEL.get(btc.get("status"), btc.get("status") or "")
    created_iso = _iso_date_only(btc.get("created_on"))
    record = {
        "Status":                     status_label,
        "Template Title":             template_title,
        "Template ID":                template_id,
        "BT Client ID":               btc.get("id") or "",
        "Client ID":                  btc.get("client_id") or client.get("id") or "",
        "Personal Code":              client.get("personal_code") or "",
        "Full Name":                  client.get("full_name") or "",
        "Phone":                      normalise_phone(client.get("phone")),
        "Email":                      client.get("email") or "",
        "Bank Account":               client.get("bank_account") or "",
        "Bank Code":                  client.get("bank_code") or "",
        "Street Address":             client.get("street_address") or "",
        "City":                       client.get("city") or "",
        "Zip Code":                   client.get("zip_code") or "",
        "Invoice Reference":          btc.get("invoice_reference") or "",
        "Date Created":               created_iso,
        "Days Since Created":         _days_since(btc.get("created_on"), today),
        "Last Scheduled Debit":       _iso_date_only(
            btc.get("subscription_billing_scheduled_on")
        ),
        # Cross-ref columns filled in by annotate_with_active_elsewhere.
        "Currently Active Elsewhere": "",
        "Active Elsewhere Templates": "",
    }
    return record


# ─── Excel writer ────────────────────────────────────────────────────────────
def write_workbook(path, dump_records, summary):
    """openpyxl: 4-sheet workbook with bold/frozen header, auto-width
    columns, right-aligned Days Since Created, and conditional fill on
    the Status cell (paused=amber, pending=red, inactive=grey).

    Each per-status sheet is sorted by Days Since Created ASC so the
    freshest leads sit at the top. Records with blank Days come last.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # The default sheet becomes the Summary tab.
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, summary)

    status_fills = {
        "Paused":   PatternFill("solid", fgColor="FFE0B2"),  # amber
        "Pending":  PatternFill("solid", fgColor="FFCDD2"),  # red
        "Inactive": PatternFill("solid", fgColor="E0E0E0"),  # grey
    }
    for label in ("Paused", "Pending", "Inactive"):
        rows = [r for r in dump_records if r["Status"] == label]
        # Sort by Days Since Created ASC, blanks (== "") last via tuple key.
        rows.sort(key=lambda r: (r["Days Since Created"] == "",
                                 r["Days Since Created"]))
        ws = wb.create_sheet(label)
        _write_data_sheet(ws, rows, status_fill=status_fills[label],
                          font=Font, alignment=Alignment,
                          col_letter=get_column_letter)

    wb.save(path)


def _write_summary_sheet(ws, summary):
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)

    ws["A1"] = "Revio Non-Active Subscriber Dump"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {datetime.now(timezone.utc).isoformat(timespec='seconds')}"

    row = 4
    ws.cell(row=row, column=1, value="Total non-active subscribers").font = bold
    ws.cell(row=row, column=2, value=summary["total"])
    row += 2

    ws.cell(row=row, column=1, value="By status").font = bold
    row += 1
    for label in ("Paused", "Pending", "Inactive"):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=summary["by_status"][label])
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Currently active elsewhere").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Total")
    ws.cell(row=row, column=2, value=summary["active_elsewhere_total"])
    row += 1
    for label in ("Paused", "Pending", "Inactive"):
        ws.cell(row=row, column=1,
                value=f"  …of which {label.lower()}")
        ws.cell(row=row, column=2,
                value=summary["active_elsewhere_by_status"][label])
        row += 1
    row += 1

    ws.cell(row=row, column=1,
            value="Top 10 templates by non-active subscriber count").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Template").font = bold
    ws.cell(row=row, column=2, value="Count").font = bold
    row += 1
    for title, count in summary["top_templates"]:
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=2, value=count)
        row += 1

    ws.column_dimensions[get_column_letter(1)].width = 60
    ws.column_dimensions[get_column_letter(2)].width = 16
    ws.cell(row=4, column=2).alignment = Alignment(horizontal="right")


def _write_data_sheet(ws, rows, *, status_fill, font, alignment, col_letter):
    bold = font(bold=True)
    right = alignment(horizontal="right")

    # Header row
    for ci, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = bold
    ws.freeze_panes = "A2"

    # Data rows
    for ri, rec in enumerate(rows, start=2):
        for ci, name in enumerate(COLUMNS, start=1):
            v = rec.get(name, "")
            cell = ws.cell(row=ri, column=ci, value=v)
            if name == "Status":
                cell.fill = status_fill
            elif name == "Days Since Created" and isinstance(v, int):
                cell.alignment = right

    # Auto-width — capped so a long street_address doesn't blow out the
    # whole sheet. The +2 is breathing room for the bold header glyphs.
    for ci, name in enumerate(COLUMNS, start=1):
        max_len = len(name)
        for rec in rows:
            v = rec.get(name, "")
            s = "" if v is None else str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter(ci)].width = min(max_len + 2, 50)


# ─── DUMP_AUDIT row append ───────────────────────────────────────────────────
def _get_sheets_service():
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds_json = os.environ.get("GOOGLE_SHEETS_CREDENTIALS")
    if not creds_json:
        raise RuntimeError("GOOGLE_SHEETS_CREDENTIALS not set")
    creds = Credentials.from_service_account_info(
        json.loads(creds_json), scopes=SHEETS_SCOPES,
    )
    return build("sheets", "v4", credentials=creds)


def _ensure_audit_tab(svc):
    """Create DUMP_AUDIT tab + write the header row if the tab does not
    already exist. No-op if it does. Returns the tab name."""
    meta = svc.spreadsheets().get(
        spreadsheetId=SHEET_ID,
        fields="sheets(properties(title))",
    ).execute()
    titles = {s["properties"]["title"] for s in meta.get("sheets", [])}
    if AUDIT_TAB in titles:
        return AUDIT_TAB

    svc.spreadsheets().batchUpdate(
        spreadsheetId=SHEET_ID,
        body={"requests": [{"addSheet": {"properties": {"title": AUDIT_TAB}}}]},
    ).execute()
    svc.spreadsheets().values().update(
        spreadsheetId=SHEET_ID,
        range=f"{AUDIT_TAB}!A1",
        valueInputOption="USER_ENTERED",
        body={"values": [AUDIT_HEADER]},
    ).execute()
    logger.info("Created %s tab + header row in master sheet", AUDIT_TAB)
    return AUDIT_TAB


def append_audit_row(*, run_ts_iso, mode, total_examined,
                     paused_count, pending_count, inactive_count,
                     active_elsewhere_count, triggered_by, filename):
    """Append a single audit row to DUMP_AUDIT. Failures are logged but
    do not abort the run — the dump itself + email are the deliverable;
    the audit row is best-effort persistence."""
    try:
        svc = _get_sheets_service()
        _ensure_audit_tab(svc)
        svc.spreadsheets().values().append(
            spreadsheetId=SHEET_ID,
            range=f"{AUDIT_TAB}!A1",
            valueInputOption="USER_ENTERED",
            insertDataOption="INSERT_ROWS",
            body={"values": [[
                run_ts_iso, mode, total_examined,
                paused_count, pending_count, inactive_count,
                active_elsewhere_count, triggered_by, filename,
            ]]},
        ).execute()
        logger.info("Appended DUMP_AUDIT row to master sheet")
    except Exception as e:
        logger.error("Failed to append DUMP_AUDIT row: %s", e)


# ─── Email ───────────────────────────────────────────────────────────────────
def _build_summary_email(*, run_date, summary, dry_run, error_summary,
                         duration_seconds, output_filename):
    total = summary["total"]
    by_s = summary["by_status"]
    ae = summary["active_elsewhere_total"]

    if error_summary:
        outcome = "failure"
        headline = "Error during non-active dump"
        para = (f"Dump failed before completion. {error_summary}. Manual "
                f"investigation needed.")
    elif total == 0:
        outcome = "noop"
        headline = "No non-active subscribers found"
        para = ("No subscribers in the paused / pending / inactive buckets "
                "across any subscription template. Nothing to dump.")
    else:
        outcome = "success"
        headline = f"{total} non-active subscriber(s) dumped"
        para = (
            f"Found {total} non-active subscriber(s) — "
            f"{by_s['Paused']} paused, {by_s['Pending']} pending, "
            f"{by_s['Inactive']} inactive. "
            f"{ae} of these have the same phone number on a currently "
            f"active subscription elsewhere (do not cold-call without "
            f"checking the 'Active Elsewhere Templates' column). "
            f"Excel attached. POPIA: this file contains personal data — "
            f"handle per Project Help's POPIA policy. Do not forward "
            f"externally."
        )

    numbers = {
        "Total non-active":            total,
        "Paused":                      by_s["Paused"],
        "Pending":                     by_s["Pending"],
        "Inactive":                    by_s["Inactive"],
        "Currently active elsewhere":  ae,
    }

    next_steps = []
    top5_paused = summary.get("top_templates_by_paused") or []
    if top5_paused:
        bullets = [f"{title}: {count}" for title, count in top5_paused]
        next_steps.append("Top 5 templates by paused count: "
                          + " · ".join(bullets))
    if dry_run:
        next_steps.append(
            "Dry-run: production distribution is jd / tashlyn / tsakane. "
            "Re-run with DRY_RUN=false to email all three."
        )

    return RunSummary(
        workflow_name="Revio — Non-Active Subscriber Dump",
        run_date=run_date,
        mode="dry_run" if dry_run else "production",
        outcome=outcome,
        headline=headline,
        summary_paragraph=para,
        numbers=numbers,
        duration_seconds=duration_seconds,
        next_steps=next_steps,
        attachments_note=f"Excel workbook attached: {output_filename}",
    )


def send_summary_email(run_date, *, summary, dry_run, error_summary="",
                       duration_seconds=0.0, attachment_path=None):
    sender = os.environ.get("EMAIL_SENDER")
    pwd    = os.environ.get("EMAIL_PASSWORD")
    if not sender or not pwd:
        logger.warning("EMAIL_SENDER / EMAIL_PASSWORD not set — skipping email")
        return
    if dry_run:
        recipients = [DRY_RUN_RECIPIENT]
    else:
        recipients = list(PRODUCTION_RECIPIENTS)
    if not recipients:
        logger.warning("No recipients — skipping email")
        return

    output_filename = (Path(attachment_path).name
                       if attachment_path else "(no file)")
    rs = _build_summary_email(
        run_date=run_date, summary=summary, dry_run=dry_run,
        error_summary=error_summary, duration_seconds=duration_seconds,
        output_filename=output_filename,
    )
    subject, html_body = build_run_summary_email(rs)

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"]    = sender
    msg["To"]      = ", ".join(recipients)
    msg.set_content("HTML email — see HTML part for the run summary.")
    msg.add_alternative(html_body, subtype="html")

    if attachment_path and Path(attachment_path).exists():
        ctype, _ = mimetypes.guess_type(attachment_path)
        if not ctype:
            ctype = ("application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet")
        maintype, subtype = ctype.split("/", 1)
        with open(attachment_path, "rb") as f:
            msg.add_attachment(
                f.read(), maintype=maintype, subtype=subtype,
                filename=Path(attachment_path).name,
            )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(sender, pwd)
        smtp.send_message(msg)
    logger.info("Summary email sent → %s%s",
                ", ".join(recipients), " (DRY RUN)" if dry_run else "")


# ─── Main ────────────────────────────────────────────────────────────────────
def _triggered_by():
    """Best-effort actor label for the audit row. GitHub Actions sets
    GITHUB_ACTOR; falls back to "manual" for ad-hoc local runs."""
    return os.environ.get("GITHUB_ACTOR") or "manual"


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )
    run_date = datetime.now(timezone.utc)
    today    = date.today()
    started  = time.monotonic()
    brand_scope = os.environ.get("REVIO_BRAND_ID", "")
    logger.info("=" * 60)
    logger.info("REVIO — NON-ACTIVE SUBSCRIBER DUMP")
    logger.info("Run date    : %s", run_date.isoformat(timespec="seconds"))
    logger.info("Dry run     : %s", DRY_RUN)
    logger.info("Brand scope : %s", brand_scope or "(all brands)")
    logger.info("Triggered by: %s", _triggered_by())
    logger.info("=" * 60)

    output_path = Path("/tmp") / (
        f"revio_non_active_dump_{today.isoformat()}.xlsx"
    )
    summary_data = {
        "total": 0,
        "by_status": {"Paused": 0, "Pending": 0, "Inactive": 0},
        "by_template": [],
        "top_templates": [],
        "active_elsewhere_total": 0,
        "active_elsewhere_by_status":
            {"Paused": 0, "Pending": 0, "Inactive": 0},
    }

    try:
        templates = list_subscription_templates()
        logger.info("Found %d subscription template(s)", len(templates))

        # ── 1st pass: enumerate BTCs, bucket by status ───────────────────────
        # active_btc_pairs: list of (template_title, btc) — phone fetched
        # in the next step.
        # non_active_btc_pairs: list of (template_title, template_id, btc, bucket)
        active_btc_pairs        = []
        non_active_btc_pairs    = []
        skipped_unknown         = 0

        for tmpl in templates:
            tid    = tmpl.get("id") or tmpl.get("uuid")
            ttitle = tmpl.get("title") or tmpl.get("name") or str(tid)
            if not tid:
                logger.warning("Template missing id; skipping: %s", tmpl)
                continue
            btcs = list_clients_for_template(tid)
            t_active = t_paused = t_pending = t_inactive = t_unknown = 0
            for btc in btcs:
                bucket = classify_non_active(btc)
                if bucket == BUCKET_SKIP_ACTIVE:
                    active_btc_pairs.append((ttitle, btc))
                    t_active += 1
                elif bucket == BUCKET_SKIP_UNKNOWN:
                    skipped_unknown += 1
                    t_unknown += 1
                    logger.info(
                        "skip[unknown_status] template=%s btc_id=%s status=%r",
                        tid, btc.get("id"), btc.get("status"),
                    )
                else:
                    non_active_btc_pairs.append((ttitle, tid, btc, bucket))
                    if bucket   == BUCKET_PAUSED:   t_paused   += 1
                    elif bucket == BUCKET_PENDING:  t_pending  += 1
                    elif bucket == BUCKET_INACTIVE: t_inactive += 1
            logger.info(
                "Template %s (%s): total=%d active=%d paused=%d pending=%d "
                "inactive=%d unknown=%d",
                tid, ttitle, len(btcs),
                t_active, t_paused, t_pending, t_inactive, t_unknown,
            )

        logger.info(
            "Pass 1 complete: %d active BTCs, %d non-active BTCs, "
            "%d unknown-status skipped",
            len(active_btc_pairs), len(non_active_btc_pairs), skipped_unknown,
        )

        # ── 2nd pass: fetch Client records (cached by client_id) ─────────────
        # The same Client can appear on multiple templates (one per BTC),
        # so caching collapses duplicates. We fetch for both active and
        # non-active because the active phone index needs the phone field.
        client_cache = {}
        all_client_ids = (
            {b.get("client_id") for _, b in active_btc_pairs if b.get("client_id")}
            | {b.get("client_id") for _, _, b, _ in non_active_btc_pairs
               if b.get("client_id")}
        )
        logger.info("Fetching %d unique Client record(s)", len(all_client_ids))
        for n, cid in enumerate(all_client_ids, 1):
            fetch_client(cid, client_cache)
            if n % 250 == 0:
                logger.info("  …fetched %d / %d", n, len(all_client_ids))

        # ── Build active phone index ─────────────────────────────────────────
        active_records = []
        for ttitle, btc in active_btc_pairs:
            client = client_cache.get(btc.get("client_id"))
            if not client:
                continue
            active_records.append({
                "phone":          normalise_phone(client.get("phone")),
                "template_title": ttitle,
            })
        active_phone_index = build_active_phone_index(active_records)
        logger.info(
            "Active phone index: %d unique phone(s) across %d active BTC(s)",
            len(active_phone_index), len(active_records),
        )

        # ── Build dump records + annotate ────────────────────────────────────
        dump_records = []
        for ttitle, tid, btc, bucket in non_active_btc_pairs:
            client = client_cache.get(btc.get("client_id"))
            rec = build_dump_record(btc, client, ttitle, tid, today)
            annotate_with_active_elsewhere(rec, active_phone_index)
            dump_records.append(rec)

            # PII-masked debug line so the workflow log is auditable
            # without leaking full ID / bank values.
            logger.debug(
                "  dump[%s] template=%s personal_code=%s phone=%s "
                "active_elsewhere=%s",
                bucket, ttitle,
                _mask_pii(rec["Personal Code"]),
                _mask_pii(rec["Phone"]),
                rec["Currently Active Elsewhere"],
            )

        # ── Summary ──────────────────────────────────────────────────────────
        summary_data = build_summary_sheet_data(dump_records)
        logger.info(
            "Dump assembled: total=%d paused=%d pending=%d inactive=%d "
            "active_elsewhere=%d",
            summary_data["total"],
            summary_data["by_status"]["Paused"],
            summary_data["by_status"]["Pending"],
            summary_data["by_status"]["Inactive"],
            summary_data["active_elsewhere_total"],
        )

        # ── Write workbook ───────────────────────────────────────────────────
        write_workbook(output_path, dump_records, summary_data)
        logger.info("Wrote workbook → %s (%d row(s))",
                    output_path, len(dump_records))

        # ── Audit row ────────────────────────────────────────────────────────
        append_audit_row(
            run_ts_iso=run_date.isoformat(timespec="seconds"),
            mode="dry_run" if DRY_RUN else "production",
            total_examined=summary_data["total"],
            paused_count=summary_data["by_status"]["Paused"],
            pending_count=summary_data["by_status"]["Pending"],
            inactive_count=summary_data["by_status"]["Inactive"],
            active_elsewhere_count=summary_data["active_elsewhere_total"],
            triggered_by=_triggered_by(),
            filename=output_path.name,
        )

        # ── Email ────────────────────────────────────────────────────────────
        send_summary_email(
            run_date=run_date,
            summary=summary_data,
            dry_run=DRY_RUN,
            duration_seconds=time.monotonic() - started,
            attachment_path=str(output_path),
        )

    except Exception as e:
        logger.exception("Dump failed: %s", e)
        try:
            send_summary_email(
                run_date=run_date,
                summary=summary_data,
                dry_run=DRY_RUN,
                error_summary=str(e),
                duration_seconds=time.monotonic() - started,
                attachment_path=str(output_path) if output_path.exists() else None,
            )
        except Exception as e2:
            logger.error("Summary email also failed: %s", e2)
        sys.exit(1)

    logger.info("=" * 60)
    logger.info("DONE")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
